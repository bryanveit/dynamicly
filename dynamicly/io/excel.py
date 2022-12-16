import numpy as np
import pandas as pd
from dynamicly.io.loadmat import loadmat
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import load_workbook
import pathlib

def mat_to_excel(path_in=None, path_out=None, column=None, index='Frequency'):
    if path_in is None:
        root = tk.Tk()
        root.withdraw()
        path_in = filedialog.askopenfilename()
    if path_out is None:
        path_out = path_in.replace('.mat', '.xlsx')

    data = loadmat(path_in,key=False)
    writer = pd.ExcelWriter(path_out, engine='openpyxl')
    if os.path.exists(path_out):
        book = load_workbook(path_out)
        writer.book = book
    for key, val in data.items():
        data_numpy = val
        if column is not None:
            data_pandas = pd.DataFrame(data_numpy[:, -1],
                                       index=data_numpy[:, 0],
                                       columns=[column])
        else:
            data_pandas = pd.DataFrame(data_numpy[:, -1],
                                       index=data_numpy[:, 0],
                                       columns=[key])
        data_pandas.to_excel(writer, sheet_name=key,
                             index_label=index)
        writer.save()
    writer.close()
    return

    # if len(data.keys()) == 1:
    #     the_key = list(data.keys())[0]
    #     data_numpy = data[the_key]
    #     data_pandas = pd.DataFrame(data_numpy[:, -1], index=data_numpy[:, 0],
    #                                columns=[column])
    #     data_pandas.to_excel(path_out, index_label=index)
    # else:
    #     return


def to_excel(data, path_out=None, column=None, index='Frequency'):
    if 'xlsx' not in path_out.split('.')[-1]:
        if len(path_out.split('.')) == 1:
            path_out = path_out + '.xlsx'
        else:
            path_out = path_out.split('.')[0] + '.xlsx'

    if path_out is None:
        return
    if isinstance(data, dict):

        writer = pd.ExcelWriter(path_out, engine='openpyxl')
        if os.path.exists(path_out):
            book = load_workbook(path)
            writer.book = book
        for key, val in data.items():
            data_numpy = val
            if column is not None:
                data_pandas = pd.DataFrame(data_numpy[:, -1],
                                           index=data_numpy[:, 0],
                                           columns=[column])
            else:
                data_pandas = pd.DataFrame(data_numpy[:, -1],
                                           index=data_numpy[:, 0],
                                           columns=[key])
            data_pandas.to_excel(writer, sheet_name=key,
                                 index_label=index)
            writer.save()
        writer.close()
    else:
        data_numpy = data
        data_pandas = pd.DataFrame(data_numpy[:, -1], index=data_numpy[:, 0],
                                   columns=[column])
        data_pandas.to_excel(path_out, index_label=index)
    return

def to_csv(data, path_out=None, column=None, index='Frequency (Hz)'):
    pass
    if 'csv' not in path_out.split('.')[-1]:
        if len(path_out.split('.')) == 1:
            path_out = path_out + '.csv'
        else:
            path_out = path_out.split('.')[0] + '.csv'

    if path_out is None:
        return
    if isinstance(data, dict):
        pass
        # Needs work
        # for key, val in data.items():
        #     data_numpy = val
        #     data_pandas = pd.DataFrame(data_numpy[:, -1],
        #                                index=data_numpy[:, 0],
        #                                columns=[key])
        #     data_pandas.to_csv(path_out, index_label=index)

    else:
        data_numpy = data
        data_pandas = pd.DataFrame(data_numpy[:, -1], index=data_numpy[:, 0],
                                   columns=[column])
        data_pandas.to_csv(path_out, index_label=index)
    return

